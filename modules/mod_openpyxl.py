from openpyxl import load_workbook  #Чтение файла
from openpyxl import Workbook       #Запись файла

import os

'''
dir = os.path.dirname(__file__)
# wb_1 = load_workbook(dir + "\\Excel\\test.xlsx")
'''

wb_read = load_workbook("./Excel/test.xlsx")
wb_write = Workbook()
